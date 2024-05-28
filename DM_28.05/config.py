from openpyxl.styles import PatternFill

excel_data = None
excel_path = None

x1, y1, x2, y2 = 300, 127, 602, 244
threshold_value = 128
sharpness_value = 50

green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")