import pytesseract
import re
from openpyxl import load_workbook
import pandas as pd
from config import excel_data, excel_path, green_fill, yellow_fill

def extract_digits(text):
    digits = re.sub(r'\D', '', text)
    print(f"Extracted digits: {digits}")
    return digits

def apply_ocr(img_bw_threshold):
    return pytesseract.image_to_string(img_bw_threshold, lang='rus')

def compare_with_excel_data(text):
    if excel_data is not None:
        results = []
        workbook = load_workbook(excel_path)
        sheet = workbook.active

        filtered_text = extract_digits(text)
        print(f"Filtered text: {filtered_text}")

        for i, row in excel_data.iterrows():
            row_match = False
            for j, cell in enumerate(row):
                if pd.isna(cell):
                    continue
                cell = str(cell)
                filtered_cell = extract_digits(cell)
                print(f"Comparing: {filtered_text} with {filtered_cell}")
                if filtered_text == filtered_cell:
                    sheet.cell(row=i+2, column=j+1).fill = green_fill
                    results.append((i, cell, "Exact"))
                    row_match = True
                elif filtered_text in filtered_cell or filtered_cell in filtered_text:  # Исправлено "или" на "or"
                    sheet.cell(row=i+2, column=j+1).fill = yellow_fill
                    results.append((i, cell, "Partial"))
                    row_match = True

        workbook.save(excel_path)

        if not results:
            results.append(("Ошибка", "Номер не найден"))
        return results
    return []