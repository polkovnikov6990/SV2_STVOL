import cv2
import tkinter as tk
from PIL import Image, ImageTk
import os
import numpy as np
import pytesseract
import tkinter.filedialog
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re

# Глобальные переменные для хранения данных из Excel
excel_data = None
excel_path = None

# Цвета для выделения совпадений
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Функция для загрузки файла Excel
def load_excel_data():
    global excel_data, excel_path
    excel_path = tkinter.filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
    if excel_path:
        excel_data = pd.read_excel(excel_path)
        print("Данные из Excel загружены")
        show_excel_data()
    return excel_data

# Функция для отображения загруженных данных в интерфейсе
def show_excel_data():
    if excel_data is not None:
        data_str = excel_data.to_string(index=False)
        text_box.delete(1.0, tk.END)
        text_box.insert(tk.END, data_str)

# Функция для извлечения цифр из строки
def extract_digits(text):
    digits = re.sub(r'\D', '', text)
    print(f"Extracted digits: {digits}")  # Отладочный вывод
    return digits

# Функция для сравнения распознанного текста с данными из Excel и выделения совпадений
def compare_with_excel_data(text):
    if excel_data is not None:
        results = []
        workbook = load_workbook(excel_path)
        sheet = workbook.active

        filtered_text = extract_digits(text)
        print(f"Filtered text: {filtered_text}")  # Отладочный вывод

        for i, row in excel_data.iterrows():
            row_match = False
            for j, cell in enumerate(row):
                if pd.isna(cell):
                    continue
                cell = str(cell)
                filtered_cell = extract_digits(cell)
                print(f"Comparing: {filtered_text} with {filtered_cell}")  # Отладочный вывод
                if filtered_text == filtered_cell:
                    sheet.cell(row=i+2, column=j+1).fill = green_fill
                    results.append((i, cell, "Exact"))
                    row_match = True
                elif filtered_text in filtered_cell or filtered_cell in filtered_text:
                    sheet.cell(row=i+2, column=j+1).fill = yellow_fill
                    results.append((i, cell, "Partial"))
                    row_match = True

        workbook.save(excel_path)

        if not results:
            results.append(("Ошибка", "Номер не найден"))
        return results
    return []

# Глобальные переменные для координат прямоугольника
x1, y1, x2, y2 = 300, 127, 602, 244
# Глобальная переменная для порогового значения
threshold_value = 128
# Глобальная переменная для резкости
sharpness_value = 50

def set_sharpness(cap, sharpness_value):
    # Check if the camera supports the sharpness property
    if not cap.set(cv2.CAP_PROP_SHARPNESS, sharpness_value):
        print("Не удалось установить резкость камеры")

def update_threshold(val):
    global threshold_value
    threshold_value = int(val)
    update_captured_image()  # Обновляем изображение с новым порогом

def update_image():
    ret, frame = cap.read()
    if ret:
        global x1, y1, x2, y2
        cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 255, 0), 2)
        cv_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        img = Image.fromarray(cv_image)
        img = img.resize((600, int(600 * frame.shape[0] / frame.shape[1])), Image.Resampling.LANCZOS)
        imgtk = ImageTk.PhotoImage(image=img)
        video_label.imgtk = imgtk
        video_label.configure(image=imgtk)
        video_label.after(10, update_image)

def update_x1(val):
    global x1
    x1 = int(val)

def update_y1(val):
    global y1
    y1 = int(val)

def update_x2(val):
    global x2
    x2 = int(val)

def update_y2(val):
    global y2
    y2 = int(val)

def capture_image():
    ret, frame = cap.read()
    if ret:
        global x1, y1, x2, y2
        cropped_frame = frame[y1:y2, x1:x2]
        cv2.imwrite('captured_frame.jpg', cropped_frame)
        update_captured_image()

def update_captured_image():
    global threshold_value
    img_bw_cv = None
    if os.path.exists('captured_frame.jpg'):
        img = Image.open('captured_frame.jpg')
        imgtk = ImageTk.PhotoImage(image=img)
        capture_label.imgtk = imgtk
        capture_label.configure(image=imgtk)

        # Преобразование в черно-белое
        img_bw = img.convert('L')
        img_bw_cv = np.array(img_bw)  # Преобразование из PIL Image в NumPy массив для OpenCV
        img_bw_path = 'captured_frame_bw.jpg'
        img_bw.save(img_bw_path)

        # Применение пороговой обработки
        _, img_bw_threshold = cv2.threshold(img_bw_cv, threshold_value, 255, cv2.THRESH_BINARY)
        
        # Сохранение порогового изображения
        img_bw_threshold_path = 'captured_frame_bw_threshold.jpg'
        cv2.imwrite(img_bw_threshold_path, img_bw_threshold)

        # Преобразование обратно в PIL Image и отображение
        img_bw_threshold_pil = Image.fromarray(img_bw_threshold)
        imgtk_bw_threshold = ImageTk.PhotoImage(image=img_bw_threshold_pil)
        capture_label_bw_threshold.imgtk = imgtk_bw_threshold
        capture_label_bw_threshold.configure(image=imgtk_bw_threshold)

        # Применение OCR к пороговому изображению с указанием языка
        text = pytesseract.image_to_string(img_bw_threshold, lang='rus')
        print("Считанный текст с изображения:")
        print(text)

        # Сравнение с данными из Excel
        results = compare_with_excel_data(text)
        if results and results[0][0] == "Ошибка":
            result_label.config(text=results[0][1], fg="red")
        else:
            results_text = "\n".join([f"Row {row}: {cell} ({match_type})" for row, cell, match_type in results])
            result_label.config(text=results_text, fg="black")

cap = cv2.VideoCapture(0)
if not cap.isOpened():
    print("Не удалось подключиться к камере")
    exit()

set_sharpness(cap, sharpness_value)

root = tk.Tk()
root.title("Камера")

left_frame = tk.Frame(root, width=800, height=600)
left_frame.pack(side=tk.LEFT, expand=True, fill=tk.BOTH)
center_frame = tk.Frame(root, width=800, height=600)
center_frame.pack(side=tk.LEFT, expand=True, fill=tk.BOTH)
right_frame = tk.Frame(root, width=800, height=600)
right_frame.pack(side=tk.LEFT, expand=True, fill=tk.BOTH)

load_button = tk.Button(right_frame, text="Добавить базу данных", command=load_excel_data)
load_button.pack()

video_label = tk.Label(left_frame)
video_label.pack(expand=True)
capture_label = tk.Label(center_frame)
capture_label.pack(expand=True)
capture_label_bw_threshold = tk.Label(center_frame)
capture_label_bw_threshold.pack(expand=True)

capture_button = tk.Button(right_frame, text="Захватить изображение", command=capture_image)
capture_button.pack()

scale_x1 = tk.Scale(right_frame, from_=0, to=1280, orient=tk.HORIZONTAL, label="X1", command=update_x1)
scale_x1.set(x1)
scale_x1.pack()
scale_y1 = tk.Scale(right_frame, from_=0, to=720, orient=tk.HORIZONTAL, label="Y1", command=update_y1)
scale_y1.set(y1)
scale_y1.pack()
scale_x2 = tk.Scale(right_frame, from_=0, to=1280, orient=tk.HORIZONTAL, label="X2", command=update_x2)
scale_x2.set(x2)
scale_x2.pack()
scale_y2 = tk.Scale(right_frame, from_=0, to=720, orient=tk.HORIZONTAL, label="Y2", command=update_y2)
scale_y2.set(y2)
scale_y2.pack()

threshold_scale = tk.Scale(right_frame, from_=0, to=255, orient=tk.HORIZONTAL, label="Threshold", command=update_threshold)
threshold_scale.set(threshold_value)
threshold_scale.pack()

sharpness_scale = tk.Scale(right_frame, from_=0, to=100, orient=tk.HORIZONTAL, label="Sharpness", command=lambda val: set_sharpness(cap, int(val)))
sharpness_scale.set(sharpness_value)
sharpness_scale.pack()

result_label = tk.Label(right_frame, text="", wraplength=400)
result_label.pack()

text_box = tk.Text(right_frame)

# Начать обновление изображения с камеры
update_image()

root.mainloop()
