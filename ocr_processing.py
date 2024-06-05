import tkinter as tk
from tkinter import Label, Text, Scrollbar, VERTICAL, RIGHT, Y, END
from PIL import Image, ImageTk
import pytesseract
import os
import re
from openpyxl import load_workbook
import pandas as pd
from config import excel_data, excel_path, green_fill, yellow_fill

# Убедитесь, что переменная окружения установлена
os.environ['TESSDATA_PREFIX'] = '/usr/local/share/tessdata/'

def extract_digits(text):
    digits = re.sub(r'\D', '', text)
    print(f"Extracted digits: {digits}")
    return digits

def apply_ocr(img_bw_threshold):
    return pytesseract.image_to_string(img_bw_threshold, lang='rus')

def compare_with_excel_data(text):
    if excel_data is not None:
        results = []
        workbook = load_workbook(excel_path)
        sheet = workbook.active

        filtered_text = extract_digits(text)
        print(f"Filtered text: {filtered_text}")

        for i, row in excel_data.iterrows():
            row_match = False
            for j, cell in enumerate(row):
                if pd.isna(cell):
                    continue
                cell = str(cell)
                filtered_cell = extract_digits(cell)
                print(f"Comparing: {filtered_text} with {filtered_cell}")
                if filtered_text == filtered_cell:
                    sheet.cell(row=i+2, column=j+1).fill = green_fill
                    results.append((i, cell, "Exact"))
                    row_match = True
                elif filtered_text in filtered_cell or filtered_cell in filtered_text:
                    sheet.cell(row=i+2, column=j+1).fill = yellow_fill
                    results.append((i, cell, "Partial"))
                    row_match = True

        workbook.save(excel_path)

        if not results:
            results.append(("Ошибка", "Номер не найден"))
        return results
    return []

class App:
    def __init__(self, root):
        self.root = root
        
    def show_image_window(self):
        # Создание нового окна
        new_window = tk.Toplevel(self.root)
        new_window.title("Recognized Image")
        
        # Загрузка и отображение изображения
        img = Image.open(self.img_path)
        img = img.resize((400, 400), Image.ANTIALIAS)  # Изменение размера изображения
        img_tk = ImageTk.PhotoImage(img)
        
        img_label = Label(new_window, image=img_tk)
        img_label.image = img_tk  # Сохранение ссылки на изображение
        img_label.pack()
        
        # Применение OCR и отображение текста
        text = apply_ocr(self.img_path)
        text_label = Label(new_window, text=text)
        text_label.pack()
        
        # Сравнение с данными из Excel и отображение результатов
        results = compare_with_excel_data(text)
        
        # Добавление виджета для отображения результатов
        result_text = Text(new_window, wrap='word', height=10, width=50)
        result_text.pack(side='left', fill='both', expand=True)
        
        # Добавление полосы прокрутки
        scrollbar = Scrollbar(new_window, orient=VERTICAL, command=result_text.yview)
        scrollbar.pack(side=RIGHT, fill=Y)
        result_text.config(yscrollcommand=scrollbar.set)
        
        # Вывод результатов в текстовый виджет
        for result in results:
            result_text.insert(END, f"{result}\n")

# Пример использования
root = tk.Tk()
app = App(root)
root.mainloop()